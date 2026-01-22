import streamlit as st
import pandas as pd
import os
import tempfile
from openpyxl import load_workbook
from openpyxl.pivot.table import PivotTable, PivotField, Reference
from openpyxl.pivot.cache import PivotCache, CacheDefinition

st.set_page_config(page_title="GSTR-1 Processor", layout="centered")
st.title("GSTR-1 Excel Processor")

# ---------------- Utilities ---------------- #

def normalize_columns(df):
    df.columns = (
        df.columns.astype(str)
        .str.strip()
        .str.replace(r"\s+", " ", regex=True)
    )
    return df


def find_column_by_keywords(df, keywords, label):
    for col in df.columns:
        col_l = col.lower()
        if all(k.lower() in col_l for k in keywords):
            return col
    raise KeyError(f"{label} column not found")


def get_column_letter_by_header(ws, header_name):
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=col).value == header_name:
            return ws.cell(row=1, column=col).column_letter
    raise KeyError(f"Column '{header_name}' not found in {ws.title}")

# ---------------- Session State ---------------- #

if "processed" not in st.session_state:
    st.session_state.processed = False
    st.session_state.outputs = {}

# ---------------- UI ---------------- #

company_code = st.text_input("Company Code")

sd_file = st.file_uploader("Upload SD File", type="xlsx")
sr_file = st.file_uploader("Upload SR File", type="xlsx")
tb_file = st.file_uploader("Upload TB File", type="xlsx")
gl_file = st.file_uploader("Upload GL Dump File", type="xlsx")

# ---------------- Processing ---------------- #

if st.button("Process Files"):

    if not all([company_code, sd_file, sr_file, tb_file, gl_file]):
        st.error("All inputs are mandatory")
        st.stop()

    try:
        outputs = {}

        with tempfile.TemporaryDirectory() as tmpdir:
            paths = {}
            for f, name in [
                (sd_file, "sd.xlsx"),
                (sr_file, "sr.xlsx"),
                (tb_file, "tb.xlsx"),
                (gl_file, "gl.xlsx"),
            ]:
                p = os.path.join(tmpdir, name)
                with open(p, "wb") as out:
                    out.write(f.getbuffer())
                paths[name] = p

            # ---------- DATA LOADING & MERGING ---------- #
            df_sales = pd.concat(
                [
                    normalize_columns(pd.read_excel(paths["sd.xlsx"])),
                    normalize_columns(pd.read_excel(paths["sr.xlsx"]))
                ],
                ignore_index=True
            )

            df_gl = normalize_columns(pd.read_excel(paths["gl.xlsx"]))
            gl_text = find_column_by_keywords(df_gl, ["g/l", "long", "text"], "GL Text")
            gl_acct = find_column_by_keywords(df_gl, ["g/l", "account"], "GL Account")
            val_col = find_column_by_keywords(df_gl, ["value"], "Amount")
            doc_col = find_column_by_keywords(df_gl, ["document"], "Document")

            gst_accounts = ["Central GST Payable", "Integrated GST Payable", "State GST Payable"]
            df_gst = df_gl[df_gl[gl_text].isin(gst_accounts)]
            df_rev = df_gl[df_gl[gl_acct].astype(str).str.startswith("3")]

            df_tb = normalize_columns(pd.read_excel(paths["tb.xlsx"]))
            tb_text = find_column_by_keywords(df_tb, ["g/l", "acct", "long"], "TB Text")
            debit = find_column_by_keywords(df_tb, ["period", "d"], "Debit")
            credit = find_column_by_keywords(df_tb, ["period", "c"], "Credit")

            df_tb_gst = df_tb[df_tb[tb_text].isin(gst_accounts)].copy()
            df_tb_gst["Difference as per TB"] = df_tb_gst[credit] - df_tb_gst[debit]

            summary_df = pd.merge(
                df_gst.groupby(gl_text, as_index=False)[val_col].sum()
                .rename(columns={gl_text: "GST Type", val_col: "GST Payable as per GL"}),
                df_tb_gst.groupby(tb_text, as_index=False)["Difference as per TB"].sum()
                .rename(columns={tb_text: "GST Type"}),
                on="GST Type",
                how="left"
            ).fillna(0)

            # ---------- WRITE INITIAL WORKBOOK ---------- #
            gstr_path = os.path.join(tmpdir, f"{company_code}_GSTR-1_Workbook.xlsx")
            with pd.ExcelWriter(gstr_path, engine="openpyxl") as writer:
                df_sales.to_excel(writer, "Sales register", index=False)
                df_rev.to_excel(writer, "Revenue", index=False)
                df_gst.to_excel(writer, "GST payable", index=False)
                summary_df.to_excel(writer, "GST Summary", index=False)

            # ---------- EXCEL POST PROCESS ---------- #
            wb = load_workbook(gstr_path)
            ws_sales = wb["Sales register"]
            ws_rev = wb["Revenue"]
            ws_gst = wb["GST payable"]
            ws_sum = wb["GST Summary"]

            # --- CREDIT NOTE NEGATIVES --- #
            doc_type = get_column_letter_by_header(ws_sales, "Document Type")
            amt_cols = ["Taxable value", "IGST Amt", "CGST Amt", "SGST/UTGST Amt"]
            amt_letters = [get_column_letter_by_header(ws_sales, c) for c in amt_cols]

            for r in range(2, ws_sales.max_row + 1):
                if ws_sales[f"{doc_type}{r}"].value == "C":
                    for c in amt_letters:
                        if isinstance(ws_sales[f"{c}{r}"].value, (int, float)):
                            ws_sales[f"{c}{r}"].value *= -1

            # --- SALES SUMMARY COLUMN --- #
            inv = get_column_letter_by_header(ws_sales, "Invoice type")
            tax = get_column_letter_by_header(ws_sales, "Tax rate")
            ss_col_idx = ws_sales.max_column + 1
            ws_sales.cell(1, ss_col_idx, "Sales summary")

            for r in range(2, ws_sales.max_row + 1):
                it = ws_sales[f"{inv}{r}"].value
                dt = ws_sales[f"{doc_type}{r}"].value
                tr = float(ws_sales[f"{tax}{r}"].value or 0)
                val = ""
                if it == "SEWOP": val = "SEZWOP"
                elif it == "SEWP": val = "SEWP"
                elif it not in ("SEWOP", "SEWP") and tr == 0: val = "Exempt supply"
                elif it == "B2B" and dt == "C" and tr != 0: val = "B2B Credit Notes"
                elif it == "B2B" and dt != "C" and tr != 0: val = "B2B Supplies"
                elif it == "B2CS" and tr != 0: val = "B2C Supplies"
                ws_sales.cell(r, ss_col_idx, val)

            # --- VLOOKUPS & CROSS LOOKUPS --- #
            gf8 = get_column_letter_by_header(ws_sales, "Generic Field 8")
            rev_doc_col = get_column_letter_by_header(ws_rev, "Document Number")
            gst_doc_col = get_column_letter_by_header(ws_gst, "Document Number")

            sr_last = ws_sales.max_column
            ws_sales.cell(1, sr_last + 1, "Revenue VLOOKUP")
            ws_sales.cell(1, sr_last + 2, "GST Payable VLOOKUP")

            for r in range(2, ws_sales.max_row + 1):
                ws_sales.cell(r, sr_last + 1, f'=IFERROR(VLOOKUP({gf8}{r},Revenue!{rev_doc_col}:{rev_doc_col},1,FALSE),"Not Found")')
                ws_sales.cell(r, sr_last + 2, f'=IFERROR(VLOOKUP({gf8}{r},\'GST payable\'!{gst_doc_col}:{gst_doc_col},1,FALSE),"Not Found")')

            # --- PIVOT TABLE: SALES SUMMARY --- #
            # Create a new sheet for the pivot
            ws_pivot = wb.create_sheet("Sales summary")
            
            # Define source data range (Sales register sheet)
            # Reference(worksheet, min_col, min_row, max_col, max_row)
            pivot_source_range = f"'Sales register'!$A$1:${get_column_letter_by_header(ws_sales, 'Sales summary')}${ws_sales.max_row}"
            
            # Get Column Indices for Pivot Fields (0-indexed for fields)
            headers = [cell.value for cell in ws_sales[1]]
            idx_gstin = headers.index("GSTIN of Taxpayer")
            idx_summary = headers.index("Sales summary")
            idx_taxable = headers.index("Taxable value")
            idx_igst = headers.index("IGST Amt")
            idx_cgst = headers.index("CGST Amt")
            idx_sgst = headers.index("SGST/UTGST Amt")

            # Setup Pivot Cache
            pc = PivotCache(cacheSource=CacheDefinition(Reference(ws_sales, min_col=1, min_row=1, max_col=ws_sales.max_column, max_row=ws_sales.max_row)))
            
            # Create Pivot Table Object
            pt = PivotTable(cacheDefinition=pc, name="SalesSummaryPivot")
            pt.pivotTableStyleInfo = None # Default style
            
            # 1. Filter Area: GSTIN of Taxpayer
            pt.pageFields.append(PivotField(fld=idx_gstin))
            
            # 2. Rows Area: Sales summary
            pt.rowFields.append(PivotField(fld=idx_summary))
            
            # 3. Values Area: Sum of Amount Columns
            # note: DataField refers to the indices in the order they appear in the source
            data_cols = [
                (idx_taxable, "Sum of Taxable value"),
                (idx_igst, "Sum of IGST Amt"),
                (idx_cgst, "Sum of CGST Amt"),
                (idx_sgst, "Sum of SGST/UTGST Amt")
            ]
            
            for fld_idx, name in data_cols:
                pt.dataFields.append(pd.api.types.DataField(fld=fld_idx, name=name, subtotal='sum'))

            # Add Pivot to the new sheet
            ws_pivot.add_pivot_table(pt, "A3")

            # Finalize Sum sheet formula
            ws_sum.cell(1, 4, "Net Difference")
            for r in range(2, ws_sum.max_row + 1):
                c = ws_sum.cell(r, 4, f"=B{r}+C{r}")
                c.number_format = "0.00"

            wb.save(gstr_path)

            with open(gstr_path, "rb") as f:
                outputs["GSTR-1 Workbook.xlsx"] = f.read()

        st.session_state.outputs = outputs
        st.session_state.processed = True
        st.success("Processing completed successfully")

    except Exception as e:
        st.error(f"Error: {str(e)}")

# ---------------- Download ---------------- #

if st.session_state.processed:
    for k, v in st.session_state.outputs.items():
        st.download_button(f"Download {k}", v, file_name=k)



